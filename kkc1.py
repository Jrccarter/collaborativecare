# -*- coding: utf-8 -*-
"""
Created on Sat Mar  2 17:18:44 2019

@author: carol
"""

import numpy as np
import math
#x = np.linspace(0,5,50)
y = []
list = [80,60,70,65,63,60,95,179,100,100,100,100,185,120,110,100,100,179,130,
        120,110,100,95,90,85]
for i in list:
    y.append(i)
    
myoutputfile = open('myoutput.txt','w')

#for i in range(len(x)):
#    thisline = str(x[i]) + '\t' + str(y[i])
#    print (thisline)
#    myoutputfile.write(thisline)
#    myoutputfile.write('\n')
time =0
for i in range(len(list)):
    thisline = str(time) + '\t' + str(y[i])
    print (thisline)
    myoutputfile.write(thisline)
    myoutputfile.write('\n')
    time+=100
    
from openpyxl import Workbook

outwb = Workbook()
thissheet = outwb.create_sheet(title = 'My data ')

for i in range(len(x)):
    thissheet.cell(row=i+1,column=1).value = x[i]
    thissheet.cell(row=i+1,column=2).value = y[i]

outwb.save('myexcelfile.xlsx')